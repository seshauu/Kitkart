from flask import Flask, render_template, redirect, request, flash, jsonify
import serial
import openpyxl
import os
from datetime import datetime, timedelta
import threading
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = "your_secret_key"  # For flashing messages

# Path to Excel file
EXCEL_FILE = "rfid_log.xlsx"

# Predefined UID to trolley number mapping
UID_TROLLEY_MAPPING = {
    '63a0a5f': '001',
    'b3252ce': '002',
    # Add more UID-trolley mappings here
}

# Email configuration
EMAIL_ADDRESS = 'praveentansam@gmail.com'  # Replace with your email address
EMAIL_PASSWORD = 'dkoy yloq vhsl dhrf'  # Replace with your email password
RECIPIENT_EMAILS = ['seshanjai@gmail.com','Shanmugam.Murugan@rnaipl.com'] # Replace with recipient email address

# Create Excel file if it doesn't exist
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFID Log"
        ws.append(["UID", "Entry Date", "Trolley", "Entry Time", "Exit Date", "Exit Time", "Remarks", "Due Date"])
        wb.save(EXCEL_FILE)

# Function to read UID from ESP32 via serial
def read_rfid_serial():
    try:
        ser = serial.Serial('COM3', 115200, timeout=1)  # Change 'COM3' to the correct port
        while True:
            uid = ser.readline().decode('utf-8').strip()
            if uid:
                log_rfid_entry(uid)
            time.sleep(1)  # Add a delay to avoid overwhelming the system with reads
    except Exception as e:
        print(f"Error reading serial: {e}")

# Function to log RFID entry/exit into Excel
def log_rfid_entry(uid):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if UID exists in the Excel file
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == uid:
            found = True
            entry_date = row[1].value
            exit_date = row[4].value

            # If entry exists but exit is missing, log exit time
            if entry_date and not exit_date:
                row[4].value = datetime.now().strftime("%Y-%m-%d")  # Exit date
                row[5].value = datetime.now().strftime("%H:%M:%S")  # Exit time
                row[6].value = "Service Completed"  # Remarks
                wb.save(EXCEL_FILE)  # Save after updating
                print(f"Logged exit for UID {uid} with Trolley {row[2].value}.")
            break

    # If UID is not found, log a new entry with trolley number
    if not found:
        trolley_number = UID_TROLLEY_MAPPING.get(uid, "Unknown Trolley")  # Retrieve trolley number or default to "Unknown"
        ws.append([uid, datetime.now().strftime("%Y-%m-%d"), trolley_number, datetime.now().strftime("%H:%M:%S"), None, None, None, None])
        wb.save(EXCEL_FILE)  # Save after adding a new entry
        print(f"Logged UID {uid} with Trolley {trolley_number} successfully.")

# Background task to continuously read RFID serial data
def start_rfid_reading():
    thread = threading.Thread(target=read_rfid_serial)
    thread.daemon = True  # Daemonize thread to close with the program
    thread.start()

# Function to send email notifications
def send_email(uid, trolley, days_left):
    subject = f"Due Date Notification for Trolley {uid}"
    body = f"Dear User,\n\nThe due date for {trolley} Trolley is in {days_left} days.\nPlease take necessary actions.\n\nBest regards,\nYour KITKART TEAM"

    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = ', '.join(RECIPIENT_EMAILS)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
            print(f"Email sent for UID {uid} with Trolley {trolley}.")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Function to check for due dates and send email notifications
def check_due_dates():
    while True:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        today = datetime.now().date()

        for row in ws.iter_rows(min_row=2, values_only=False):
            due_date_str = row[7].value
            if due_date_str:
                try:
                    due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    days_left = (due_date - today).days
                    uid = row[0].value
                    trolley = row[2].value

                    if days_left == 7 or days_left == 1:
                        send_email(uid, trolley, days_left)

                except ValueError:
                    continue

        time.sleep(86400)  # Check every day

# Redirect root URL to /records
@app.route('/')
def index():
    return redirect('/records')

# Route to scan RFID
@app.route('/scan', methods=['GET'])
def scan_rfid():
    return render_template('scan.html')

# Route to display and edit records
@app.route('/records', methods=['GET', 'POST'])
def edit_record():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fetch all data from Excel
    records = []
    today = datetime.now().date()

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * (8 - len(row))
        due_date_str = row[7]
        due_date = None
        if due_date_str:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            except ValueError:
                due_date = None

        days_left = (due_date - today).days if due_date else None

        records.append({
            'uid': row[0],
            'entry_date': row[1],
            'trolley': row[2],
            'entry_time': row[3],
            'exit_date': row[4],
            'exit_time': row[5],
            'concern': row[6],
            'due_date': row[7],
            'days_left': days_left if days_left is not None and days_left >= 0 else 'Overdue',
            'repair_done_by': row[8]  # Assuming this is the new column for "Repair Done By"
        })

    if request.method == 'POST':
        uid = request.form['uid']
        remarks = request.form['remarks']
        new_due_date = request.form['due_date']
        repair_done_by = request.form['repair_done_by']  # Capture repair done by name

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == uid:
                row[6].value = remarks
                row[7].value = new_due_date if not row[7].value else row[7].value
                row[8].value = repair_done_by  # Update the Repair Done By column
                flash(f"Updated record for UID {uid}.")
                wb.save(EXCEL_FILE)
                break

        return redirect('/records')

    return render_template('records.html', records=records)


# Route to return records as JSON for AJAX requests
@app.route('/get_records', methods=['GET'])
def get_records():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fetch all data from Excel
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append({
            'uid': row[0],
            'entry_date': row[1],
            'trolley': row[2],
            'entry_time': row[3],
            'exit_date': row[4],
            'exit_time': row[5],
            'concern': row[6],
            'due_date': row[7]
        })

    return jsonify(records)

if __name__ == '__main__':
    create_excel_file()
    start_rfid_reading()  # Start the background task for automatic RFID logging
    threading.Thread(target=check_due_dates, daemon=True).start()  # Start the due date checking thread
    app.run(debug=True)
