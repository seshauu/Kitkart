from flask import Flask, render_template, redirect, request, flash, jsonify
import serial
import openpyxl
import os
from datetime import datetime, timedelta
import threading
import time

app = Flask(__name__)
app.secret_key = "your_secret_key"  # For flashing messages

# Path to Excel file
EXCEL_FILE = "rfid_log.xlsx"

# Predefined UID to trolley number mapping
UID_TROLLEY_MAPPING = {
    '63a0a5f': '001',
    'b3252ce': '002',
    # Add more UID-trolley mappings here
}

# Create Excel file if it doesn't exist
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFID Log"
        ws.append(["UID", "Entry Date", "Trolley", "Entry Time", "Exit Date", "Exit Time", "Remarks", "Due Date"])
        wb.save(EXCEL_FILE)

# Function to read UID from ESP32 via serial
def read_rfid_serial():
    try:
        ser = serial.Serial('COM3', 115200, timeout=1)  # Change 'COM7' to the correct port
        while True:
            uid = ser.readline().decode('utf-8').strip()
            if uid:
                log_rfid_entry(uid)
            time.sleep(1)  # Add a delay to avoid overwhelming the system with reads
    except Exception as e:
        print(f"Error reading serial: {e}")

# Function to log RFID entry/exit into Excel
def log_rfid_entry(uid):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if UID exists in the Excel file
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == uid:
            found = True
            entry_date = row[1].value
            exit_date = row[4].value

            # If entry exists but exit is missing, log exit time
            if entry_date and not exit_date:
                row[4].value = datetime.now().strftime("%Y-%m-%d")  # Exit date
                row[5].value = datetime.now().strftime("%H:%M:%S")  # Exit time
                row[6].value = "Service Completed"  # Remarks
                # Remove the line that adds the due date
                # row[7].value = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")  # Due Date (7 days from exit)
                wb.save(EXCEL_FILE)  # Save after updating
                print(f"Logged exit for UID {uid} with Trolley {row[2].value}.")
            break

    # If UID is not found, log a new entry with trolley number
    if not found:
        trolley_number = UID_TROLLEY_MAPPING.get(uid, "Unknown Trolley")  # Retrieve trolley number or default to "Unknown"
        ws.append([uid, datetime.now().strftime("%Y-%m-%d"), trolley_number, datetime.now().strftime("%H:%M:%S"), None, None, None, None])
        wb.save(EXCEL_FILE)  # Save after adding a new entry
        print(f"Logged UID {uid} with Trolley {trolley_number} successfully.")


# Background task to continuously read RFID serial data
def start_rfid_reading():
    thread = threading.Thread(target=read_rfid_serial)
    thread.daemon = True  # Daemonize thread to close with the program
    thread.start()

# Redirect root URL to /records
@app.route('/')
def index():
    return redirect('/records')

# Route to scan RFID
@app.route('/scan', methods=['GET'])
def scan_rfid():
    return render_template('/scan.html')

# Route to display and edit records
from datetime import datetime


@app.route('/records', methods=['GET', 'POST'])
def edit_record():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fetch all data from Excel
    records = []
    today = datetime.now().date()  # Get the current date

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ensure the row has the necessary length by padding with None if needed
        row = list(row) + [None] * (8 - len(row))

        # Convert the due date from string to date if it exists
        due_date_str = row[7]
        due_date = None
        if due_date_str:
            try:
                # Assuming the date format is 'YYYY-MM-DD', modify if necessary
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            except ValueError:
                due_date = None

        # Calculate days left until due date
        days_left = (due_date - today).days if due_date else None

        records.append({
            'uid': row[0],
            'entry_date': row[1],
            'trolley': row[2],
            'entry_time': row[3],
            'exit_date': row[4],
            'exit_time': row[5],
            'concern': row[6],
            'due_date': row[7],
            'days_left': days_left if days_left is not None and days_left >= 0 else 'Overdue'
        })

    if request.method == 'POST':
        uid = request.form['uid']
        remarks = request.form['remarks']
        new_due_date = request.form['due_date']

        # Find the UID in the Excel and update the row
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == uid:
                # Update remarks
                row[6].value = remarks

                # Update due date only if it is not already set
                if not row[7].value:  # Only update if due date is None
                    row[7].value = new_due_date
                    flash(f"Updated due date for UID {uid} to {new_due_date}.")
                else:
                    flash(f"Due date for UID {uid} cannot be changed as it is already set to {row[7].value}.")

                wb.save(EXCEL_FILE)
                break

        return redirect('/records')

    return render_template('records.html', records=records)


# Route to return records as JSON for AJAX requests
@app.route('/get_records', methods=['GET'])
def get_records():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fetch all data from Excel
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append({
            'uid': row[0],
            'entry_date': row[1],
            'trolley': row[2],
            'entry_time': row[3],
            'exit_date': row[4],
            'exit_time': row[5],
            'concern': row[6],
            'due_date': row[7]
        })

    return jsonify(records)

if __name__ == '__main__':
    create_excel_file()
    start_rfid_reading()  # Start the background task for automatic RFID logging
    app.run(debug=True)