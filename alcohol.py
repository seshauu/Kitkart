import serial
import time
import openpyxl
from datetime import datetime

# Set up serial communication (ensure the correct COM port)
ser = serial.Serial('COM3', 9600, timeout=1)  # Replace 'COM3' with your Arduino port
time.sleep(2)  # Wait for Arduino to initialize

# Load the attendance workbook and select the active sheet
try:
    wb = openpyxl.load_workbook('C:/Users/pravin/Documents/attendance.xlsx')
    sheet = wb.active
except FileNotFoundError:
    print("Error: The specified Excel file was not found.")
    exit(1)

# User ID to Name mapping
user_dict = {
    '1': 'Shanmugam San',
    '3': 'Muthusivam San'
}

# Function to log attendance in Excel
def log_attendance(finger_id, name, alcohol_level):
    row = sheet.max_row + 1  # Move to the next available row
    current_time = datetime.now()  # Get the current date and time
    status = 'Present' if alcohol_level < 600 else 'Absent'  # Set attendance status based on alcohol level
    test_status = 'Status-OK' if alcohol_level < 600 else 'NG'  # Set test status

    # Fill the attendance details
    sheet.cell(row=row, column=1, value=current_time.strftime("%Y-%m-%d"))
    sheet.cell(row=row, column=2, value=current_time.strftime("%H:%M:%S"))
    sheet.cell(row=row, column=3, value=finger_id)
    sheet.cell(row=row, column=4, value=name)
    sheet.cell(row=row, column=5, value=alcohol_level)  # Log the alcohol level
    sheet.cell(row=row, column=6, value=test_status)  # Log the attendance status in column F
    sheet.cell(row=row, column=7, value=status)  # Log the alcohol test status in column G

    # Save to a new path
    try:
        wb.save('C:/Users/pravin/Documents/attendance.xlsx')
    except PermissionError:
        print("Error: Permission denied when saving the Excel file. Make sure it's not open elsewhere.")

while True:
    if ser.in_waiting > 0:  # Check if data is available from Arduino
        data = ser.readline().decode('utf-8').strip()
        print("Received from Arduino:", data)

        # Process the data based on what Arduino sends
        if "Found ID #" in data:
            # Extract fingerprint ID from the string
            finger_id = data.split("#")[1].strip()
            print(f"Fingerprint ID: {finger_id}")

            # Retrieve the name from user_dict
            name = user_dict.get(finger_id, "Unknown User")
            print(f"User Name: {name}")

        elif "Alcohol Level:" in data:
            # Extract alcohol level from the string
            alcohol_level = int(data.split(":")[1].strip())
            print(f"Alcohol Level: {alcohol_level}")

        elif data == "Status-OK":            
            # Log the attendance
            log_attendance(finger_id, name, alcohol_level)
            print(f"Attendance logged for Fingerprint ID {finger_id} ({name}) with Alcohol Level {alcohol_level}")

        elif data == "NG":
            print(f"Fingerprint ID {finger_id} ({name}) failed alcohol test (Alcohol Level {alcohol_level}).")
            # Log the attendance with Absent status
            log_attendance(finger_id, name, alcohol_level)

    time.sleep(1)  # Small delay to prevent flooding
