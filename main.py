from flask import Flask, render_template, redirect, request, flash
import serial
import openpyxl
import os
from datetime import datetime, timedelta
import threading
import time



app = Flask(__name__)
app.secret_key = "your_secret_key"  # For flashing messages

# Path to Excel file
EXCEL_FILE = "rfid_log.xlsx"

# Create Excel file if it doesn't exist
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFID Log"
        ws.append(["UID", "Entry Date", "Entry Time", "Exit Date", "Exit Time", "Remarks", "Due Date"])
        wb.save(EXCEL_FILE)

# Function to read UID from ESP32 via serial
def read_rfid_serial():
    try:
        ser = serial.Serial('COM7', 115200, timeout=1)  # Change 'COM7' to the correct port
        while True:
            uid = ser.readline().decode('utf-8').strip()
            if uid:
                log_rfid_entry(uid)
            time.sleep(1)  # Add a delay to avoid overwhelming the system with reads
    except Exception as e:
        print(f"Error reading serial: {e}")

# Function to log RFID entry/exit into Excel
def log_rfid_entry(uid):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Check if UID exists in the Excel file
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == uid:
            found = True
            entry_date = row[1].value
            exit_date = row[3].value

            # If entry exists but exit is missing, log exit time
            if entry_date and not exit_date:
                row[3].value = datetime.now().strftime("%Y-%m-%d")  # Exit date
                row[4].value = datetime.now().strftime("%H:%M:%S")  # Exit time
                row[5].value = "Service Completed"  # Remarks
                row[6].value = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")  # Due Date (7 days from exit)
                break

    # If UID is not found, log a new entry
    if not found:
        ws.append([uid, datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%H:%M:%S"), None, None, None, None])

    wb.save(EXCEL_FILE)
    print(f"Logged UID {uid} successfully.")

# Background task to continuously read RFID serial data
def start_rfid_reading():
    thread = threading.Thread(target=read_rfid_serial)
    thread.daemon = True  # Daemonize thread to close with the program
    thread.start()

# Route to scan RFID
@app.route('/scan', methods=['GET'])
def scan_rfid():
    return render_template('/scan.html')

# Route to display and edit records
@app.route('/records', methods=['GET', 'POST'])
def edit_record():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Fetch all data from Excel
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append({
            'uid': row[0],
            'entry_date': row[1],
            'entry_time': row[2],
            'exit_date': row[3],
            'exit_time': row[4],
            'concern': row[5],
            'due_date': row[6]
        })

    if request.method == 'POST':
        uid = request.form['uid']
        remarks = request.form['remarks']
        due_date = request.form['due_date']

        # Find the UID in the Excel and update the row
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == uid:
                row[5].value = remarks
                row[6].value = due_date
                wb.save(EXCEL_FILE)
                flash(f"Updated record for UID {uid} successfully.")
                break

        return redirect('/records')

    return render_template('records.html', records=records)

if __name__ == '__main__':
    create_excel_file()
    start_rfid_reading()  # Start the background task for automatic RFID logging
    app.run(debug=True)
